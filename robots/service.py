from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import io


def generate_excel_report(data: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Generate Excel-file based on the provided data.

    Args:
        data: list of dictionaries, containing information about robots.

    Returns:
        io.BytesIO: byte stream representing the generated file.
    """

    workbook = Workbook()

    # Группируем данные по модели
    grouped_data = {}
    for entry in data:
        model = entry['model']
        if model not in grouped_data:
            grouped_data[model] = []
        grouped_data[model].append(entry)

    for model, entries in grouped_data.items():
        # Создаем отдельный лист для каждой модели
        sheet = workbook.create_sheet(title=model)
        sheet.append(['Модель', 'Версия', 'Количество за неделю'])

        for entry in entries:
            sheet.append([entry['model'], entry['version'], entry['count']])

        # Настройка ширины столбцов
        for col in range(1, 4):
            column = get_column_letter(col)
            sheet.column_dimensions[column].width = 20

    # Удаляем стандартный пустой лист
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Сохраняем файл в поток байтов
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
